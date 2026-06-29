#!/usr/bin/env python3
"""
Deterministic generator for the SYNTHETIC back-test chart.

The repository must never contain real operational data (deployed frequencies,
real Stationary KAVACH IDs, real onboard counts). This script fabricates a chart
whose *structure* mirrors a real single-section allocation chart — a ~22-station
line with a realistic spread of peak onboard counts, two approach signals at most
stations, and one busy junction terminal that exceeds a single pair's 44-marker
capacity — but whose every value is invented.

It is fully deterministic (no randomness): re-running it byte-reproduces
`synthetic_chart.xlsx`, so the fixture can be regenerated and audited from source.

The chart deliberately reproduces, on synthetic numbers, the qualitative findings
of a real-section back-test:
    * most stations fit within one frequency pair,
    * the section's interference clique forces 5 frequency pairs (window = 4),
    * at least one terminal is over-capacity and must be split across pairs.

Run:  python3 make_synthetic_chart.py [out.xlsx]
"""
from __future__ import annotations
import os
import sys

# Synthetic single section: 22 stations, ids 10001..10022 (NOT real ids).
# (peak onboard units, last-stop signals) per station. Hand-chosen to give a
# realistic spread; station 10011 is a busy junction terminal (over one pair).
SYNTHETIC_STATIONS = [
    (10001, 4, 2), (10002, 6, 2), (10003, 5, 2), (10004, 7, 2),
    (10005, 6, 2), (10006, 8, 3), (10007, 5, 2), (10008, 6, 2),
    (10009, 9, 2), (10010, 7, 2), (10011, 28, 4),   # busy junction terminal
    (10012, 6, 2), (10013, 5, 2), (10014, 8, 2), (10015, 6, 2),
    (10016, 7, 2), (10017, 5, 2), (10018, 9, 3), (10019, 6, 2),
    (10020, 7, 2), (10021, 5, 2), (10022, 4, 2),
]

SHEET = "Frequency allocation"


def build_workbook():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    # Row labels match the keywords excel_io.read_chart looks for.
    ws.cell(1, 1, "Stationary Kavach ID")
    ws.cell(2, 1, "Peak nos. of Onboard Kavach Units")
    ws.cell(3, 1, "Number of Stationary Kavach Tx slots")   # legacy pre-computed
    ws.cell(4, 1, "No. of Last Stop Signals")
    # The legacy 'Number of Stationary Kavach Tx slots' row is filled with the
    # old field heuristic (ROUNDUP((peak*120 + 100)/66)) so the fixture looks
    # like a real chart, but the back-test runs with use_slot_demand=True, which
    # IGNORES this column and recomputes n_station from first principles. No
    # 'Proposed Frequency Pair' row -> the tool also allocates pairs itself, so
    # this is a pure end-to-end reproduction.
    for k, (sid, peak, sig) in enumerate(SYNTHETIC_STATIONS):
        c = 2 + k
        legacy_sta = -(-(peak * 120 + 100) // 66)            # ROUNDUP, ints only
        ws.cell(1, c, sid)
        ws.cell(2, c, peak)
        ws.cell(3, c, legacy_sta)
        ws.cell(4, c, sig)
    return wb


def main(argv):
    out = argv[1] if len(argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "synthetic_chart.xlsx")
    build_workbook().save(out)
    print(f"wrote {out} ({len(SYNTHETIC_STATIONS)} synthetic stations)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
