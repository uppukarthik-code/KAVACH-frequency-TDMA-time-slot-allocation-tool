#!/usr/bin/env python3
"""
Excel I/O for the KAVACH allocation solver.

Reads the RDSO "FREQUENCY ALLOCATION CHART (SEC-A-SEC-B)" workbook (sheet
"Frequency allocation") and extracts, per Stationary KAVACH unit:
    station id, #Stationary KAVACH Tx slots, #Loco KAVACH Tx slots (peak
    onboard), and the existing 'Proposed Frequency Pair'.

Writes a compliant-allocation workbook with three sheets:
    Allocation | Compliance | Justification

Requires openpyxl (`pip install openpyxl`).
"""
from __future__ import annotations
import warnings

from allocation_solver import (Pair, Problem, solve_compliant, allocation_table,
                               compliance_report, justify_changes, _mk_palette,
                               KAVACH_F0)

# Row-label keywords used to locate the data rows (robust to row shifts).
_LBL_ID = "stationary kavach id"
_LBL_PEAK = "peak nos"               # Peak nos. of Onboard Kavach Units  -> loco slots
_LBL_PAIR = "proposed frequency pair"
_LBL_STA = "number of stationary"    # Number of Stationary Kavach Tx slots (OUTPUT; optional)
_LBL_SIG = "last stop signal"        # No. of Last Stop Signals -> slot-demand input
_LBL_NAME = "station name"           # optional, for readable output
_LBL_CODE = "station code"           # optional
_LBL_LAT = "lattitude"               # Stationary Unit Tower Lattitude (sic, as in chart)
_LBL_LON = "longitude"               # Stationary Unit Tower Longitude


def _find_row(ws, keyword):
    kw = keyword.lower()
    for r in range(1, ws.max_row + 1):
        for c in range(1, 4):
            v = ws.cell(r, c).value
            if isinstance(v, str) and kw in v.lower():
                return r
    return None


def _int(x):
    """Coerce a cell value to int; accepts int, float, or numeric string."""
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return int(x)
    if isinstance(x, str):
        try:
            return int(float(x.strip()))
        except ValueError:
            return None
    return None


def _float(x):
    """Coerce a cell to float; None if not numeric."""
    try:
        return float(str(x).strip()) if x not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _haversine_km(lat1, lon1, lat2, lon2):
    """Great-circle distance (km) between two lat/long points."""
    import math
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def read_chart(path, sheet="Frequency allocation"):
    """Parse the allocation workbook -> dict of per-station inputs.

    Required input rows:  Stationary Kavach ID, Peak nos. of Onboard Kavach Units.
    Optional input rows:  No. of Last Stop Signals (default 2), Proposed Frequency
                          Pair, Station Name/code, Tower Latitude/Longitude.
    The 'Number of Stationary Kavach Tx slots' row is an OUTPUT and is NOT required
    on input (the tool computes it); it is read only if present (for --legacy-slots).
    When latitude+longitude are supplied for every station, an along-route chainage
    (km) is derived and returned as `positions` for the real RF-range model."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    r_id = _find_row(ws, _LBL_ID)
    r_peak = _find_row(ws, _LBL_PEAK)
    r_pair = _find_row(ws, _LBL_PAIR)
    r_sta = _find_row(ws, _LBL_STA)
    r_sig = _find_row(ws, _LBL_SIG)
    r_name = _find_row(ws, _LBL_NAME)
    r_code = _find_row(ws, _LBL_CODE)
    r_lat = _find_row(ws, _LBL_LAT)
    r_lon = _find_row(ws, _LBL_LON)
    if r_id is None:
        raise ValueError("could not locate 'Stationary Kavach ID' row")
    if r_peak is None:
        raise ValueError("chart is missing required row "
                         "'Peak nos. of Onboard Kavach Units'")

    def cell_int(r, c, field):
        if r is None:
            return 0
        raw = ws.cell(r, c).value
        v = _int(raw)
        if v is None:
            if raw not in (None, ""):
                warnings.warn(f"{field}: non-numeric value {raw!r} at column {c} "
                              f"-> treated as 0", RuntimeWarning)
            return 0
        return v

    stations, sta, loco, existing, sig = [], {}, {}, {}, {}
    names, codes, lat, lon = {}, {}, {}, {}
    seen = set()
    for c in range(1, ws.max_column + 1):
        sid = _int(ws.cell(r_id, c).value)
        if sid is None or not (10000 <= sid <= 99999):
            continue
        if sid in seen:
            raise ValueError(f"duplicate Stationary Kavach ID {sid} in chart "
                             f"(column {c}); IDs must be unique")
        seen.add(sid)
        stations.append(sid)
        sta[sid] = cell_int(r_sta, c, "Stationary Tx slots")
        loco[sid] = cell_int(r_peak, c, "Loco slots")
        sig[sid] = cell_int(r_sig, c, "Last stop signals") if r_sig else 2
        if r_pair:
            p = _int(ws.cell(r_pair, c).value)
            if p is not None:
                existing[sid] = p
        if r_name:
            names[sid] = ws.cell(r_name, c).value
        if r_code:
            codes[sid] = ws.cell(r_code, c).value
        if r_lat:
            lat[sid] = _float(ws.cell(r_lat, c).value)
        if r_lon:
            lon[sid] = _float(ws.cell(r_lon, c).value)
    if not stations:
        raise ValueError("no Stationary Kavach IDs (10000-99999) found in chart")

    # derive along-route chainage (km) from lat/long when fully supplied
    positions = None
    if stations and all(lat.get(s) is not None and lon.get(s) is not None
                        for s in stations):
        positions, cum = {}, 0.0
        positions[stations[0]] = 0.0
        for prev, cur in zip(stations, stations[1:]):
            cum += _haversine_km(lat[prev], lon[prev], lat[cur], lon[cur])
            positions[cur] = round(cum, 3)

    return {"stations": stations, "sta_slots": sta, "loco_slots": loco,
            "signals": sig, "existing_pair": existing, "names": names,
            "codes": codes, "positions": positions}


def _parse_palette_rows(rows):
    """rows: list of dicts (any-case headers). Returns (palette, f0).
    Recognised columns: pair/id, fS/fs_mhz/stn tx, fM/fm_mhz/onboard tx.
    A row whose pair cell is 'f0'/'control'/'emergency' supplies f0."""
    def pick(d, *names):
        for n in names:
            for k, v in d.items():
                if k and k.strip().lower() == n:
                    return v
        # fuzzy contains
        for n in names:
            for k, v in d.items():
                if k and n in k.strip().lower():
                    return v
        return None

    palette, f0 = [], None
    for d in rows:
        pid = pick(d, "pair", "id")
        fs = pick(d, "fs", "fs_mhz", "stn tx", "stn", "downlink")
        fm = pick(d, "fm", "fm_mhz", "onboard tx", "onboard", "uplink")
        if pid is None and fs is None:
            continue
        if isinstance(pid, str) and pid.strip().lower() in ("f0", "control", "emergency"):
            f0 = float(fs if fs not in (None, "") else fm)
            continue
        pi = _int(pid)
        if pi is None or fs in (None, "") or fm in (None, ""):
            continue
        palette.append(Pair(pi, float(fs), float(fm)))
    if not palette:
        raise ValueError("no frequency pairs parsed from palette source")
    return palette, f0


def read_palette(path):
    """Load a frequency palette (+optional f0) from a .csv or .xlsx file."""
    if path.lower().endswith(".csv"):
        import csv
        with open(path, newline="") as fh:
            rows = list(csv.DictReader(fh))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = next((wb[s] for s in wb.sheetnames
                   if any(k in s.lower() for k in ("palette", "freq", "channel"))),
                  wb[wb.sheetnames[0]])
        hdr = [str(c.value).strip() if c.value is not None else ""
               for c in ws[1]]
        rows = [{hdr[i]: r[i] for i in range(len(hdr))}
                for r in ws.iter_rows(min_row=2, values_only=True)]
    return _parse_palette_rows(rows)


def build_problem(chart, palette=None, f0=KAVACH_F0, use_slot_demand=False,
                  peak_load_cap=None, **kw):
    """Build a Problem from a parsed chart.
    use_slot_demand=True -> compute (n_station, n_loco) with the spec-traceable
    slot_demand calculator from the chart's peak-onboard + last-stop-signals,
    instead of reading the chart's pre-computed slot columns.
    peak_load_cap -> clamp supervised trains per station to this many (Railway
    Board operational cap, RB letter 31.07.2023; e.g. 24 for the final phase)."""
    palette = palette or _mk_palette()
    if use_slot_demand:
        import slot_demand as SD
        sta, loco = {}, {}
        for sid in chart["stations"]:
            N = chart["loco_slots"].get(sid, 0)               # peak onboard units
            S = chart.get("signals", {}).get(sid, 2)          # last stop signals
            r = SD.slot_demand(SD.StationDemandInputs(peak_locos=N,
                                                      last_stop_signals=S,
                                                      peak_load_cap=peak_load_cap))
            sta[sid], loco[sid] = r["n_station"], r["n_loco"]
    else:
        sta, loco = chart["sta_slots"], chart["loco_slots"]
    return Problem(chart["stations"], sta, loco, palette, f0=f0, **kw)


def write_compliant_xlsx(path, prob: Problem, result: dict, existing_pair: dict,
                         provenance: list = None):
    import openpyxl
    wb = openpyxl.Workbook()

    # --- Allocation sheet ---
    ws = wb.active
    ws.title = "Allocation"
    rows = allocation_table(prob, result)
    headers = list(rows[0].keys()) + ["existing_pair", "changed"]
    ws.append(headers)
    for r in rows:
        old = existing_pair.get(r["station_id"])
        line = list(r.values()) + [old, "YES" if old != r["pair"] else ""]
        ws.append(line)
    ws.append([])
    ws.append([f"control/emergency f0 = {prob.f0} MHz (emergency slots P47-P70)"])
    ws.append([f"spectrum used = {result['spectrum']} pairs: "
               f"{sorted(result['used_pairs'])}",
               f"IM3: {result.get('im3_note', '')}"])

    # --- Compliance sheet ---
    wc = wb.create_sheet("Compliance")
    wc.append(["Clause / check", "Result"])
    for clause, status in compliance_report(prob, result):
        wc.append([clause, status])

    # --- Justification sheet ---
    wj = wb.create_sheet("Justification")
    wj.append(["Station", "Old pair", "New pair", "Reason"])
    for s, old, new, reason in justify_changes(prob, result, existing_pair):
        wj.append([s, old, new, reason])

    # --- Provenance sheet (audit: ties this output to code + input) ---
    if provenance:
        wp = wb.create_sheet("Provenance")
        wp.append(["Key", "Value"])
        for k, v in provenance:
            wp.append([k, v])

    wb.save(path)
    return path


def run(input_path, output_path="allocation_compliant.xlsx", **prob_kw):
    """End-to-end: read chart -> solve compliant -> write workbook + return summary."""
    chart = read_chart(input_path)
    prob = build_problem(chart, **prob_kw)
    result = solve_compliant(prob)
    write_compliant_xlsx(output_path, prob, result, chart["existing_pair"])
    return prob, result, chart


if __name__ == "__main__":
    import sys
    inp = sys.argv[1] if len(sys.argv) > 1 else None
    if not inp:
        print("usage: python3 excel_io.py <input_chart.xlsx> [output.xlsx]")
        raise SystemExit(1)
    out = sys.argv[2] if len(sys.argv) > 2 else "allocation_compliant.xlsx"
    prob, result, chart = run(inp, out)
    print(f"read {len(chart['stations'])} stations from {inp}")
    print(f"spectrum = {result['spectrum']} pairs {sorted(result['used_pairs'])}"
          f" | IM3: {result.get('im3_note')}")
    print(f"validation: {'PASS' if not result['errors'] else result['errors']}")
    print(f"written -> {out}  (Allocation / Compliance / Justification sheets)")
